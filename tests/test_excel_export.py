import json
from pathlib import Path

import openpyxl
import pytest

from rentalanalysis.calculator import analyze_property
from rentalanalysis.excel_export import build_workbook
from rentalanalysis.models import PropertyListing


def make_result(listing_data: dict, config):
    listing = PropertyListing.model_validate(listing_data)
    return analyze_property(listing, config)


def test_build_workbook_creates_file(tmp_path, sample_listing, sample_config):
    result = analyze_property(sample_listing, sample_config)
    out = tmp_path / "test.xlsx"
    build_workbook([result], out)
    assert out.exists()


def test_overview_sheet_present(tmp_path, sample_listing, sample_config):
    result = analyze_property(sample_listing, sample_config)
    out = tmp_path / "test.xlsx"
    build_workbook([result], out)
    wb = openpyxl.load_workbook(out)
    assert "Overview" in wb.sheetnames


def test_property_sheet_per_result(tmp_path, sample_listing, sample_config):
    r1 = analyze_property(sample_listing, sample_config)
    listing2_data = {
        "url": "https://portal.onehome.com/listings/99999",
        "address": "456 Oak Ave, Austin TX 78702",
        "list_price": 350000,
        "beds": 2,
        "baths": 1,
        "annual_taxes": 5000,
        "estimated_rent_monthly": 2200,
    }
    r2 = make_result(listing2_data, sample_config)
    out = tmp_path / "test.xlsx"
    build_workbook([r1, r2], out)
    wb = openpyxl.load_workbook(out)
    # Overview + 2 property sheets
    assert len(wb.sheetnames) == 3
    assert "Overview" == wb.sheetnames[0]


def test_overview_frozen_panes(tmp_path, sample_listing, sample_config):
    result = analyze_property(sample_listing, sample_config)
    out = tmp_path / "test.xlsx"
    build_workbook([result], out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Overview"]
    assert ws.freeze_panes == "A4"


def test_overview_has_address_header(tmp_path, sample_listing, sample_config):
    result = analyze_property(sample_listing, sample_config)
    out = tmp_path / "test.xlsx"
    build_workbook([result], out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Overview"]
    # Row 3, column B should be the address
    assert ws.cell(row=3, column=2).value == sample_listing.address


def test_duplicate_addresses_get_unique_sheet_names(tmp_path, sample_listing, sample_config):
    # Two listings whose addresses share the same first 31 chars must not collide.
    long_addr = "1000 Really Long Boulevard Avenue Extension, Springfield"
    d1 = {"url": "https://x/1", "address": long_addr, "list_price": 300000,
          "beds": 3, "baths": 2, "estimated_rent_monthly": 2000}
    d2 = {"url": "https://x/2", "address": long_addr, "list_price": 310000,
          "beds": 3, "baths": 2, "estimated_rent_monthly": 2100}
    r1 = make_result(d1, sample_config)
    r2 = make_result(d2, sample_config)
    out = tmp_path / "dup.xlsx"
    build_workbook([r1, r2], out, sample_config)
    wb = openpyxl.load_workbook(out)
    # Overview + 2 distinct property sheets
    assert len(wb.sheetnames) == 3
    assert len(set(wb.sheetnames)) == 3
    assert all(len(n) <= 31 for n in wb.sheetnames)


def test_incomplete_listing_gets_banner(tmp_path, sample_config):
    listing_data = {"url": "https://x/9", "address": "No Income Ln", "list_price": 300000,
                    "beds": 3, "baths": 2}
    listing = PropertyListing.model_validate(listing_data)
    result = analyze_property(listing, sample_config, use_fallback_rent=True)
    out = tmp_path / "inc.xlsx"
    build_workbook([result], out, sample_config)
    wb = openpyxl.load_workbook(out)
    ws = wb[wb.sheetnames[1]]
    assert "INCOMPLETE" in str(ws.cell(row=2, column=1).value)


def test_overview_values_reference_property_sheets(tmp_path, sample_listing, sample_config):
    result = analyze_property(sample_listing, sample_config)
    out = tmp_path / "ref.xlsx"
    build_workbook([result], out, sample_config)
    wb = openpyxl.load_workbook(out)
    ws = wb["Overview"]
    sheet = [n for n in wb.sheetnames if n != "Overview"][0]
    # At least one Overview data cell should be a cross-sheet formula to the property sheet
    found_ref = False
    for row in ws.iter_rows(min_col=2, max_col=2):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("=") and sheet[:10] in cell.value:
                found_ref = True
    assert found_ref


def test_overview_sorted_by_cash_on_cash(tmp_path, sample_config):
    # Two properties with different returns; the higher CoC must be the first column.
    good = {"url": "https://x/good", "address": "Good Deal St", "list_price": 200000,
            "beds": 3, "baths": 2, "annual_taxes": 2000, "estimated_rent_monthly": 3000}
    weak = {"url": "https://x/weak", "address": "Weak Deal Ave", "list_price": 600000,
            "beds": 3, "baths": 2, "annual_taxes": 9000, "estimated_rent_monthly": 2000}
    r_good = make_result(good, sample_config)
    r_weak = make_result(weak, sample_config)
    assert r_good.cash_on_cash > r_weak.cash_on_cash
    out = tmp_path / "sorted.xlsx"
    build_workbook([r_weak, r_good], out, sample_config)  # deliberately weak-first input
    wb = openpyxl.load_workbook(out)
    ws = wb["Overview"]
    assert ws.cell(row=3, column=2).value == "Good Deal St"   # highest CoC first


def test_ratio_formulas_are_divzero_guarded(tmp_path, sample_listing, sample_config):
    result = analyze_property(sample_listing, sample_config)
    out = tmp_path / "guard.xlsx"
    build_workbook([result], out, sample_config)
    wb = openpyxl.load_workbook(out)
    ws = wb[[n for n in wb.sheetnames if n != "Overview"][0]]
    formulas = [c.value for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=")]
    # Cap rate, DSCR and CoC all divide — each must be IFERROR-wrapped.
    guarded = [f for f in formulas if "IFERROR" in f]
    assert len(guarded) >= 3


def test_currency_formatting_present(tmp_path, sample_listing, sample_config):
    result = analyze_property(sample_listing, sample_config)
    out = tmp_path / "test.xlsx"
    build_workbook([result], out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Overview"]
    found_currency = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.number_format and "$" in cell.number_format:
                found_currency = True
                break
    assert found_currency
