from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .calculator import compute_amortization_schedule, evaluate_deal
from .models import AnalysisConfig, AnalysisResult, TargetConfig

# ── Number formats ──────────────────────────────────────────────────────────
CURRENCY_FMT = '"$"#,##0.00'
CURRENCY0_FMT = '"$"#,##0'
PERCENT_FMT = "0.00%"
NUMBER_FMT = "#,##0.0"

# ── Fill colours ─────────────────────────────────────────────────────────────
GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF6C8")   # editable inputs
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
SUBHEAD_FILL = PatternFill(fill_type="solid", fgColor="2E75B6")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D6E4F0")
CALC_FILL = PatternFill(fill_type="solid", fgColor="EBF5FB")     # derived formula cells
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")

# ── Fonts ────────────────────────────────────────────────────────────────────
WHITE_BOLD = Font(bold=True, color="FFFFFF")
NOTE_ONEHOME = Font(color="1E8449", italic=True, size=9)
NOTE_ASSUMED = Font(color="808080", italic=True, size=9)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Small styling helpers ────────────────────────────────────────────────────

def _block_header(ws, row: int, col_start: int, col_end: int, text: str) -> None:
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)


def _label(ws, row: int, col: int, text: str, bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    if bold:
        cell.font = Font(bold=True)


def _input(ws, row: int, col: int, value, fmt: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = YELLOW_FILL
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt


def _calc(ws, row: int, col: int, formula: str, fmt: str | None = None, bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=formula)
    cell.fill = CALC_FILL
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if bold:
        cell.font = Font(bold=True)


def _static(ws, row: int, col: int, value, fmt: str | None = None, bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if bold:
        cell.font = Font(bold=True)


def _note(ws, row: int, col: int, source: str) -> None:
    cell = ws.cell(row=row, column=col, value=source)
    cell.font = NOTE_ONEHOME if source == "OneHome" else NOTE_ASSUMED


# OneHome dt/dd keys already surfaced elsewhere on the sheet (normalized, colon-stripped).
_SHOWN_DETAIL_KEYS = {
    "type", "property type", "year built", "list price", "total units",
    "gross income", "net operating income", "operating expense",
    "annual taxes", "price per sq ft", "price per sq ft.",
    "hoa fee", "hoa", "association fee",
    # rent-roll fields (rendered in the Unit Rent Roll block)
    "unit #", "monthly rent", "beds", "baths", "units of this type",
    "total rent", "description",
}


def _collect_misc_listing_fields(result: AnalysisResult) -> list[tuple[str, str]]:
    """Every OneHome dt/dd field not already shown elsewhere → Miscellaneous section.

    '* Expense' fields are skipped here because they are rendered as expense lines.
    """
    out: list[tuple[str, str]] = []
    seen: set[str] = set()
    for label, value in (result.listing.listing_details or {}).items():
        key = label.strip().rstrip(":").lower()
        if value in (None, "", "--"):
            continue
        if key in _SHOWN_DETAIL_KEYS or key.endswith("expense") or key in seen:
            continue
        seen.add(key)
        out.append((label.strip().rstrip(":"), value))
    return out


# ── Per-property IRE proforma sheet ──────────────────────────────────────────

def write_property_sheet(wb: Workbook, result: AnalysisResult, targets: TargetConfig) -> None:
    title = result.listing.slug or f"Property_{id(result)}"
    ws = wb.create_sheet(title)

    L = result.listing
    price = L.list_price
    units = L.total_units or 1
    gross = result.gross_rental_income or 0.0
    monthly_total = gross / 12 if gross else 0.0

    # Derived percentages / per-unit bases for live formulas
    vacancy_pct = round(result.vacancy_loss / gross, 4) if gross else 0.05
    mgmt_pct = round(result.mgmt_fee_annual / gross, 4) if gross else 0.10
    maint_pct = round(result.maintenance_annual / gross, 4) if gross else 0.05
    capex_pct = round(result.capex_reserve / gross, 4) if gross else 0.05
    leasing_pct = round(result.leasing_fee_annual / gross, 4) if gross else 0.0
    insurance_per_unit = round(result.insurance_annual / units, 2) if units else result.insurance_annual
    recycle_per_unit = round(result.recycle_annual / units, 2) if units else 0.0
    rate = _derive_rate(result)
    closing = round(result.total_cash_invested - price * (1 - result.ltv), 2)

    # Column widths
    widths = {"A": 30, "B": 8, "C": 15, "D": 6, "E": 2,
              "F": 24, "G": 15, "H": 12, "I": 2,
              "J": 28, "K": 15, "L": 13, "M": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title
    t = ws.cell(row=1, column=1, value=f"IRE Proforma — {L.address}")
    t.font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells("A1:D1")

    # ════════════════════════════════════════════════════════════════════════
    #  RIGHT-TOP: LOAN & VALUE  (cols F/G, rows 3-10)  — anchors used elsewhere
    # ════════════════════════════════════════════════════════════════════════
    _block_header(ws, 3, 6, 8, "LOAN & VALUE")
    R_PRICE, R_LTV, R_LOAN, R_DOWN, R_CLOSE, R_REPAIR, R_INIT = 4, 5, 6, 7, 8, 9, 10
    _label(ws, R_PRICE, 6, "Purchase Price");        _input(ws, R_PRICE, 7, price, CURRENCY0_FMT)
    _label(ws, R_LTV, 6, "LTV");                     _input(ws, R_LTV, 7, round(result.ltv, 4), PERCENT_FMT)
    _label(ws, R_LOAN, 6, "Loan Amount");            _calc(ws, R_LOAN, 7, f"=G{R_PRICE}*G{R_LTV}", CURRENCY0_FMT)
    _label(ws, R_DOWN, 6, "Down Payment");           _calc(ws, R_DOWN, 7, f"=G{R_PRICE}-G{R_LOAN}", CURRENCY0_FMT)
    _label(ws, R_CLOSE, 6, "Closing Costs");         _input(ws, R_CLOSE, 7, closing, CURRENCY0_FMT)
    _label(ws, R_REPAIR, 6, "Repairs / Rehab");      _input(ws, R_REPAIR, 7, 0, CURRENCY0_FMT)
    _label(ws, R_INIT, 6, "Initial Investment", bold=True)
    _calc(ws, R_INIT, 7, f"=G{R_DOWN}+G{R_CLOSE}+G{R_REPAIR}", CURRENCY0_FMT, bold=True)

    # ════════════════════════════════════════════════════════════════════════
    #  RIGHT-MID: UNIT RENT ROLL  (cols F/G, rows 12+)
    # ════════════════════════════════════════════════════════════════════════
    _block_header(ws, 12, 6, 8, "UNIT RENT ROLL")
    roll = list(result.listing.rent_roll or [])
    rent_is_onehome = result.rent_source.startswith("OneHome")
    R_NUNITS = 13
    _label(ws, R_NUNITS, 6, "Number of Units")
    _input(ws, R_NUNITS, 7, units, NUMBER_FMT)
    _note(ws, R_NUNITS, 8, "OneHome" if result.listing.total_units else "Assumed")
    R_UNIT_HDR = 14
    ws.cell(row=R_UNIT_HDR, column=6, value="Unit").font = Font(bold=True)
    ws.cell(row=R_UNIT_HDR, column=7, value="Monthly Rent").font = Font(bold=True)
    ws.cell(row=R_UNIT_HDR, column=8, value="Source / Notes").font = Font(bold=True)

    unit_start = R_UNIT_HDR + 1
    if roll:
        display = roll[:12]
        for i, u in enumerate(display):
            r = unit_start + i
            beds = f" ({u.beds}bd)" if u.beds and u.beds not in ("--", None) else ""
            ws.cell(row=r, column=6, value=f"Unit {u.unit or i + 1}{beds}")
            _input(ws, r, 7, u.monthly_rent or 0, CURRENCY_FMT)
            note = "OneHome"
            if u.description and u.description not in ("--", None):
                note = f"OneHome — {u.description[:24]}"
            nc = ws.cell(row=r, column=8, value=note)
            nc.font = NOTE_ONEHOME
        n_rows = len(display)
    else:
        n_rows = min(units, 12)
        rent_per_unit = round(monthly_total / units, 2) if units else 0.0
        for i in range(n_rows):
            r = unit_start + i
            ws.cell(row=r, column=6, value=f"Unit {chr(ord('A') + i)}")
            _input(ws, r, 7, rent_per_unit, CURRENCY_FMT)
            nc = ws.cell(row=r, column=8, value="Assumed — gross split evenly")
            nc.font = NOTE_ASSUMED

    unit_end = unit_start + n_rows - 1
    R_RENT_TOTAL = unit_end + 1
    _label(ws, R_RENT_TOTAL, 6, "Total Monthly Rent", bold=True)
    _calc(ws, R_RENT_TOTAL, 7, f"=SUM(G{unit_start}:G{unit_end})", CURRENCY_FMT, bold=True)
    _note(ws, R_RENT_TOTAL, 8, "OneHome" if rent_is_onehome else "Assumed")
    if roll and len(roll) > n_rows:
        ws.cell(row=R_RENT_TOTAL + 1, column=6,
                value=f"(showing {n_rows} of {len(roll)} rent-roll lines)").font = NOTE_ASSUMED

    # ════════════════════════════════════════════════════════════════════════
    #  LEFT: INCOME  (cols A/B/C/D) — conservative: gross = MIN(rent roll, listed)
    # ════════════════════════════════════════════════════════════════════════
    _block_header(ws, 3, 1, 4, "GROSS INCOME")
    ir = 4
    R_LEASES = ir
    _label(ws, ir, 1, "Leases (Annualized Rent)")
    _calc(ws, ir, 3, f"=G{R_RENT_TOTAL}*12", CURRENCY_FMT)
    _note(ws, ir, 4, "OneHome" if result.rent_source.startswith("OneHome") else "Assumed")
    ir += 1
    R_PARK = ir
    _label(ws, ir, 1, "Parking / Laundry / Other")
    _input(ws, ir, 3, 0, CURRENCY_FMT)
    ir += 1
    R_OTHER = ir
    _label(ws, ir, 1, "Other Income")
    _input(ws, ir, 3, 0, CURRENCY_FMT)
    ir += 1
    R_RENTROLL = ir
    _label(ws, ir, 1, "Rent Roll Subtotal")
    _calc(ws, ir, 3, f"=SUM(C{R_LEASES}:C{R_OTHER})", CURRENCY_FMT)
    ir += 1

    listed_gi = result.listing.gross_income_annual_listed
    if listed_gi is not None:
        R_LISTGI = ir
        _label(ws, ir, 1, "Listed Gross Income (OneHome)")
        _static(ws, ir, 3, listed_gi, CURRENCY_FMT)
        _note(ws, ir, 4, "OneHome")
        ir += 1
        R_GROSS = ir
        _label(ws, ir, 1, "Gross Operating Income (used)", bold=True)
        _calc(ws, ir, 3, f"=MIN(C{R_RENTROLL},C{R_LISTGI})", CURRENCY_FMT, bold=True)
        _note(ws, ir, 4, "min — conservative")
    else:
        R_GROSS = ir
        _label(ws, ir, 1, "Gross Operating Income", bold=True)
        _calc(ws, ir, 3, f"=C{R_RENTROLL}", CURRENCY_FMT, bold=True)
    ws.cell(row=R_GROSS, column=3).fill = TOTAL_FILL
    ir += 1

    R_VAC = ir
    _label(ws, ir, 1, "Vacancy & Credit Loss")
    _input(ws, ir, 2, vacancy_pct, PERCENT_FMT)
    _calc(ws, ir, 3, f'=IF(D{R_VAC}="y",C{R_GROSS}*B{R_VAC},0)', CURRENCY_FMT)
    _input(ws, ir, 4, "y")
    ir += 1
    R_EGI = ir
    _label(ws, ir, 1, "Effective Gross Income", bold=True)
    _calc(ws, ir, 3, f"=C{R_GROSS}-C{R_VAC}", CURRENCY_FMT, bold=True)
    ws.cell(row=R_EGI, column=3).fill = TOTAL_FILL

    # ════════════════════════════════════════════════════════════════════════
    #  LEFT: OPERATING EXPENSES  (each line: B=pct, C=value, D=y/n toggle)
    #  Laid out with a running row counter so listed lines (Electric, misc) slot in.
    # ════════════════════════════════════════════════════════════════════════
    exp_hdr = R_EGI + 2
    _block_header(ws, exp_hdr, 1, 4, "OPERATING EXPENSES  (B = %, D = on?)")
    er = exp_hdr + 1  # running expense row

    def _exp_static(label, value, source):
        nonlocal er
        _label(ws, er, 1, label)
        _static(ws, er, 3, value, CURRENCY_FMT)
        _note(ws, er, 4, source)
        er += 1

    def _exp_pct(label, pct, gross_ref, on=True):
        nonlocal er
        _label(ws, er, 1, label)
        _input(ws, er, 2, pct, PERCENT_FMT)
        _calc(ws, er, 3, f'=IF(D{er}="y",C{gross_ref}*B{er},0)', CURRENCY_FMT)
        _input(ws, er, 4, "y" if on else "n")
        er += 1

    def _exp_formula(label, formula, source):
        nonlocal er
        _label(ws, er, 1, label)
        _calc(ws, er, 3, formula, CURRENCY_FMT)
        _note(ws, er, 4, source)
        er += 1

    def _exp_input(label, value):
        nonlocal er
        _label(ws, er, 1, label)
        _input(ws, er, 3, value, CURRENCY_FMT)
        er += 1

    exp_first = er
    _exp_static("Real Estate Taxes", result.taxes_annual, "OneHome" if L.annual_taxes else "Assumed")
    if L.insurance_annual_listed:
        _exp_static("Insurance", result.insurance_annual, "OneHome")
    else:
        _exp_formula("Insurance", f"={insurance_per_unit}*G{R_NUNITS}", "Assumed")
    _exp_pct("Management Fee", mgmt_pct, R_GROSS, on=True)
    if L.maintenance_annual_listed:
        _exp_static("Maintenance / Repairs", result.maintenance_annual, "OneHome")
    else:
        _exp_pct("Maintenance / Repairs", maint_pct, R_GROSS, on=True)
    _exp_static("HOA / Association", result.hoa_annual, "OneHome" if L.hoa_monthly else "Assumed")
    if result.electric_annual:
        _exp_static("Electric", result.electric_annual, "OneHome")
    _exp_input("Utilities (landlord-paid)", result.utilities_annual)
    _exp_input("Trash Removal", result.trash_annual)
    _exp_input("Water", result.water_annual)
    _exp_input("Sewer", result.sewer_annual)
    _exp_formula("Recycle", f"={recycle_per_unit}*G{R_NUNITS}", "Assumed")
    _exp_pct("Leasing Fees", leasing_pct, R_GROSS, on=(leasing_pct > 0))
    for mlabel, mval in result.misc_expense_items.items():
        _exp_static(mlabel, mval, "OneHome")
    exp_last = er - 1

    # Itemized subtotal, listed reference, and the conservative "used" total (MAX).
    R_ITEMSUM = er
    _label(ws, R_ITEMSUM, 1, "Itemized Expenses Subtotal")
    _calc(ws, R_ITEMSUM, 3, f"=SUM(C{exp_first}:C{exp_last})", CURRENCY_FMT)
    er += 1
    listed_opex = result.listed_operating_expenses
    if listed_opex is not None:
        R_LISTOPEX = er
        _label(ws, R_LISTOPEX, 1, "Listed Operating Expense (OneHome)")
        _static(ws, R_LISTOPEX, 3, listed_opex, CURRENCY_FMT)
        _note(ws, R_LISTOPEX, 4, "OneHome")
        er += 1
        R_TOTOPEX = er
        _label(ws, R_TOTOPEX, 1, "Total Operating Expenses (used)", bold=True)
        _calc(ws, R_TOTOPEX, 3, f"=MAX(C{R_ITEMSUM},C{R_LISTOPEX})", CURRENCY_FMT, bold=True)
        _note(ws, R_TOTOPEX, 4, "max — conservative")
    else:
        R_TOTOPEX = er
        _label(ws, R_TOTOPEX, 1, "Total Operating Expenses", bold=True)
        _calc(ws, R_TOTOPEX, 3, f"=C{R_ITEMSUM}", CURRENCY_FMT, bold=True)
    ws.cell(row=R_TOTOPEX, column=3).fill = TOTAL_FILL
    er += 1
    R_CAPEX = er
    _label(ws, R_CAPEX, 1, "Replacement Reserves (CapEx)")
    _input(ws, R_CAPEX, 2, capex_pct, PERCENT_FMT)
    _calc(ws, R_CAPEX, 3, f'=IF(D{R_CAPEX}="y",C{R_GROSS}*B{R_CAPEX},0)', CURRENCY_FMT)
    _input(ws, R_CAPEX, 4, "y")
    er += 1
    R_TOTNET = er
    _label(ws, R_TOTNET, 1, "Total Net Operating Expenses", bold=True)
    _calc(ws, R_TOTNET, 3, f"=C{R_TOTOPEX}+C{R_CAPEX}", CURRENCY_FMT, bold=True)
    ws.cell(row=R_TOTNET, column=3).fill = TOTAL_FILL
    er += 2

    # ── NOI (two-tier) ──
    R_NOI = er
    _label(ws, R_NOI, 1, "Net Operating Income (NOI)", bold=True)
    _calc(ws, R_NOI, 3, f"=C{R_EGI}-C{R_TOTOPEX}", CURRENCY_FMT, bold=True)
    ws.cell(row=R_NOI, column=3).fill = GREEN_FILL
    ws.cell(row=R_NOI, column=4, value="cap rate").font = NOTE_ASSUMED
    er += 1
    R_ENOI = er
    _label(ws, R_ENOI, 1, "Effective NOI (after CapEx)", bold=True)
    _calc(ws, R_ENOI, 3, f"=C{R_EGI}-C{R_TOTNET}", CURRENCY_FMT, bold=True)
    ws.cell(row=R_ENOI, column=3).fill = GREEN_FILL
    ws.cell(row=R_ENOI, column=4, value="DSCR").font = NOTE_ASSUMED
    er += 2

    # ── DEBT SERVICE ──
    _block_header(ws, er, 1, 4, "DEBT SERVICE")
    R_DLOAN, R_DRATE, R_DAMORT = er + 1, er + 2, er + 3
    R_DADS, R_DMDS, R_DSCR, R_MAXDS = er + 4, er + 5, er + 6, er + 7
    _label(ws, R_DLOAN, 1, "Loan Amount")
    _calc(ws, R_DLOAN, 3, f"=G{R_LOAN}", CURRENCY_FMT)
    _label(ws, R_DRATE, 1, "Interest Rate")
    _input(ws, R_DRATE, 3, round(rate, 5), PERCENT_FMT)
    _label(ws, R_DAMORT, 1, "Amortization (months)")
    _input(ws, R_DAMORT, 3, 360, NUMBER_FMT)
    _label(ws, R_DADS, 1, "Annual Debt Service", bold=True)
    _calc(ws, R_DADS, 3, f"=PMT(C{R_DRATE}/12,C{R_DAMORT},-C{R_DLOAN})*12", CURRENCY_FMT, bold=True)
    _label(ws, R_DMDS, 1, "Monthly Debt Service")
    _calc(ws, R_DMDS, 3, f"=C{R_DADS}/12", CURRENCY_FMT)
    _label(ws, R_DSCR, 1, "DSCR (Eff. NOI / ADS)", bold=True)
    _calc(ws, R_DSCR, 3, f"=C{R_ENOI}/C{R_DADS}", NUMBER_FMT, bold=True)
    _label(ws, R_MAXDS, 1, f"Max Debt Svc @ {targets.min_dscr:g}x DSCR")
    _calc(ws, R_MAXDS, 3, f"=C{R_ENOI}/{targets.min_dscr}", CURRENCY_FMT)

    # ════════════════════════════════════════════════════════════════════════
    #  RIGHT: VALUATION & RETURNS  (cols J/K, rows 3-15)
    # ════════════════════════════════════════════════════════════════════════
    _block_header(ws, 3, 10, 11, "VALUATION & RETURNS")
    rows = [
        ("Annual Gross Rent",      f"=C{R_GROSS}", CURRENCY_FMT),
        ("Vacancy & Credit Loss",  f"=C{R_VAC}",    CURRENCY_FMT),
        ("Effective Gross Income", f"=C{R_EGI}",    CURRENCY_FMT),
        ("Operating Expenses",     f"=C{R_TOTOPEX}", CURRENCY_FMT),
        ("Net Operating Income",   f"=C{R_NOI}",    CURRENCY_FMT),
    ]
    r = 4
    for label, formula, fmt in rows:
        _label(ws, r, 10, label)
        _calc(ws, r, 11, formula, fmt)
        r += 1
    R_CAP, R_VAL, R_UCF, R_LCF, R_COC = r, r + 1, r + 2, r + 3, r + 4
    _label(ws, R_CAP, 10, "Cap Rate", bold=True)
    _calc(ws, R_CAP, 11, f"=C{R_NOI}/G{R_PRICE}", PERCENT_FMT, bold=True)
    ws.cell(row=R_CAP, column=11).fill = GREEN_FILL
    _label(ws, R_VAL, 10, "Valuation @ Cap Rate")
    _calc(ws, R_VAL, 11, f"=IF(K{R_CAP}>0,C{R_NOI}/K{R_CAP},0)", CURRENCY0_FMT)
    _label(ws, R_UCF, 10, "Unleveraged Cash Flow")
    _calc(ws, R_UCF, 11, f"=C{R_NOI}-C{R_CAPEX}", CURRENCY_FMT)
    _label(ws, R_LCF, 10, "Leveraged Cash Flow", bold=True)
    _calc(ws, R_LCF, 11, f"=C{R_ENOI}-C{R_DADS}", CURRENCY_FMT, bold=True)
    _label(ws, R_COC, 10, "Cash-on-Cash", bold=True)
    _calc(ws, R_COC, 11, f"=K{R_LCF}/G{R_INIT}", PERCENT_FMT, bold=True)
    ws.cell(row=R_COC, column=11).fill = GREEN_FILL

    # ════════════════════════════════════════════════════════════════════════
    #  RIGHT: PURCHASE DECISION DASHBOARD  (cols J/K/L/M)
    # ════════════════════════════════════════════════════════════════════════
    dash_start = R_COC + 2
    _block_header(ws, dash_start, 10, 13, "PURCHASE DECISION DASHBOARD")
    hdr = dash_start + 1
    ws.cell(row=hdr, column=10, value="Metric").font = Font(bold=True, color="1F4E79")
    ws.cell(row=hdr, column=11, value="Actual").font = Font(bold=True, color="1F4E79")
    ws.cell(row=hdr, column=12, value="Target").font = Font(bold=True, color="1F4E79", italic=True)
    ws.cell(row=hdr, column=13, value="Verdict").font = Font(bold=True, color="1F4E79")

    # (label, actual_formula, target, fmt, op)
    metrics = [
        ("Monthly Cash Flow",      f"=K{R_LCF}/12",  targets.min_monthly_cash_flow, CURRENCY_FMT, ">="),
        ("Annual Cash-on-Cash",    f"=K{R_COC}",     targets.min_cash_on_cash_pct,  PERCENT_FMT,  ">="),
        ("Cap Rate",               f"=K{R_CAP}",     targets.min_cap_rate_pct,      PERCENT_FMT,  ">="),
        ("DSCR",                   f"=C{R_DSCR}",    targets.min_dscr,              NUMBER_FMT,   ">="),
        ("LTV (lower better)",     f"=G{R_LTV}",     targets.max_ltv,               PERCENT_FMT,  "<="),
    ]
    deal = evaluate_deal(result, targets)
    check_vals = list(deal["checks"].values())
    metric_rows = []
    r = hdr + 1
    for i, (label, actual, target, fmt, op) in enumerate(metrics):
        ws.cell(row=r, column=10, value=label)
        ac = ws.cell(row=r, column=11, value=actual); ac.number_format = fmt; ac.font = Font(bold=True)
        ac.border = BORDER
        tc = ws.cell(row=r, column=12, value=target); tc.number_format = fmt; tc.fill = YELLOW_FILL
        tc.font = Font(italic=True); tc.border = BORDER
        comp = ">=" if op == ">=" else "<="
        vf = f'=IF(K{r}{comp}L{r},"✓ PASS","✗ FAIL")'
        vc = ws.cell(row=r, column=13, value=vf); vc.font = Font(bold=True); vc.border = BORDER
        vc.fill = GREEN_FILL if (i < len(check_vals) and check_vals[i]) else RED_FILL
        metric_rows.append(r)
        r += 1

    # Deal verdict
    cf_row, coc_row = metric_rows[0], metric_rows[1]
    verdict_formula = (
        f'=IF(M{cf_row}="✓ PASS",'
        f'IF(M{coc_row}="✓ PASS","🟢 GO — TAKE THE DEAL","🟡 BORDERLINE — REVIEW"),'
        f'IF(M{coc_row}="✓ PASS","🟡 BORDERLINE — REVIEW","🔴 NO-GO — PASS"))'
    )
    ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=13)
    vcell = ws.cell(row=r, column=10, value=verdict_formula)
    vcell.font = Font(bold=True, size=13, color="FFFFFF")
    vcell.alignment = Alignment(horizontal="center", vertical="center")
    verdict_color = {"GO": "1E8449", "BORDERLINE": "9A7D0A", "NO-GO": "922B21"}[deal["verdict"]]
    vcell.fill = PatternFill(fill_type="solid", fgColor=verdict_color)
    ws.row_dimensions[r].height = 26
    r += 2

    # ════════════════════════════════════════════════════════════════════════
    #  RIGHT: LISTING (OneHome exact) vs CALCULATED
    # ════════════════════════════════════════════════════════════════════════
    _block_header(ws, r, 10, 13, "LISTING (OneHome exact) vs CALCULATED")
    r += 1
    ws.cell(row=r, column=10, value="Metric").font = Font(bold=True, color="1F4E79")
    ws.cell(row=r, column=11, value="Listing").font = Font(bold=True, color="1F4E79")
    ws.cell(row=r, column=12, value="Calculated").font = Font(bold=True, color="1F4E79")
    ws.cell(row=r, column=13, value="Variance").font = Font(bold=True, color="1F4E79")
    r += 1

    # (label, listed_value, calc_formula)
    comp_rows = [
        ("Gross Income", result.listed_gross_income, f"=C{R_RENTROLL}"),
        ("Operating Expense", result.listed_operating_expenses, f"=C{R_ITEMSUM}"),
        ("Net Operating Income", result.listed_noi, f"=C{R_NOI}"),
    ]
    opex_calc_cell = None
    opex_listed_cell = None
    for label, listed_val, calc_formula in comp_rows:
        _label(ws, r, 10, label, bold=(label == "Net Operating Income"))
        if listed_val is not None:
            _static(ws, r, 11, listed_val, CURRENCY_FMT)
        else:
            nc = ws.cell(row=r, column=11, value="N/A"); nc.font = NOTE_ASSUMED
        _calc(ws, r, 12, calc_formula, CURRENCY_FMT)
        if listed_val is not None:
            _calc(ws, r, 13, f"=L{r}-K{r}", CURRENCY_FMT)
        if label == "Operating Expense":
            opex_calc_cell, opex_listed_cell = f"L{r}", f"K{r}"
        r += 1

    # OpEx source + assessment
    if result.listed_operating_expenses is not None and opex_calc_cell:
        src = result.listed_opex_source or ""
        _label(ws, r, 10, "OpEx source")
        sc = ws.cell(row=r, column=11, value=src); sc.font = NOTE_ONEHOME
        ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=13)
        r += 1
        _label(ws, r, 10, "OpEx Assessment", bold=True)
        flag_formula = (
            f'=IF({opex_calc_cell}>{opex_listed_cell},'
            f'"⚠ OVER — listing understates expenses","✓ OK — within listing")'
        )
        fcell = ws.cell(row=r, column=11, value=flag_formula)
        ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=13)
        fcell.font = Font(bold=True)
        fcell.fill = RED_FILL if result.opex_exceeds_listed else GREEN_FILL
        r += 1

    # ════════════════════════════════════════════════════════════════════════
    #  MISCELLANEOUS — every other OneHome listing field (nothing dropped)
    # ════════════════════════════════════════════════════════════════════════
    misc_fields = _collect_misc_listing_fields(result)
    if misc_fields:
        r += 1
        _block_header(ws, r, 10, 13, "MISCELLANEOUS — OneHome Listing Fields")
        r += 1
        for label, value in misc_fields:
            ws.cell(row=r, column=10, value=label).font = Font(size=9)
            vc = ws.cell(row=r, column=11, value=value)
            vc.font = Font(size=9)
            ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=13)
            _note(ws, r, 14, "OneHome")
            r += 1

    # ════════════════════════════════════════════════════════════════════════
    #  AMORTIZATION SNAPSHOTS  (left, below debt service)
    # ════════════════════════════════════════════════════════════════════════
    amort_start = R_MAXDS + 3
    _block_header(ws, amort_start, 1, 4, "AMORTIZATION SNAPSHOTS")
    hr = amort_start + 1
    ws.cell(row=hr, column=1, value="Year").font = Font(bold=True)
    ws.cell(row=hr, column=2, value="Cum. Principal").font = Font(bold=True)
    ws.cell(row=hr, column=3, value="Cum. Interest").font = Font(bold=True)
    ws.cell(row=hr, column=4, value="Balance").font = Font(bold=True)
    ws.column_dimensions["B"].width = 15
    rr = hr + 1
    if result.loan_amount > 0 and result.monthly_payment > 0:
        schedule = _amort_from_payment(result.loan_amount, result.monthly_payment, 30)
        for year in [1, 5, 10, 15, 20, 25, 30]:
            m_end = year * 12
            if m_end > len(schedule):
                break
            sub = schedule[:m_end]
            _static(ws, rr, 1, f"Year {year}")
            _static(ws, rr, 2, round(sum(m["principal"] for m in sub), 2), CURRENCY0_FMT)
            _static(ws, rr, 3, round(sum(m["interest"] for m in sub), 2), CURRENCY0_FMT)
            _static(ws, rr, 4, sub[-1]["balance"], CURRENCY0_FMT)
            rr += 1
    else:
        ws.cell(row=rr, column=1, value="N/A — loan amount not available")


def _derive_rate(result: AnalysisResult) -> float:
    """Back-derive annual interest rate from monthly payment and loan amount."""
    la, m = result.loan_amount, result.monthly_payment
    n = 30 * 12
    if la <= 0 or m <= 0:
        return 0.07
    lo, hi = 0.0, 0.5
    for _ in range(60):
        rate = (lo + hi) / 2
        r = rate / 12
        computed = la * (r * (1 + r) ** n) / ((1 + r) ** n - 1) if rate > 0 else la / n
        if computed < m:
            lo = rate
        else:
            hi = rate
    return (lo + hi) / 2


def _amort_from_payment(principal: float, monthly_payment: float, term_years: int) -> list[dict]:
    n = term_years * 12
    lo, hi = 0.0, 0.02
    mid = 0.0
    for _ in range(50):
        mid = (lo + hi) / 2
        computed = principal / n if mid == 0 else principal * (mid * (1 + mid) ** n) / ((1 + mid) ** n - 1)
        if computed < monthly_payment:
            lo = mid
        else:
            hi = mid
    return compute_amortization_schedule(principal, mid * 12, term_years)


# ── Overview / Comparison Sheet ───────────────────────────────────────────────

def write_comparison_sheet(wb: Workbook, results: list[AnalysisResult], targets: TargetConfig) -> None:
    ws = wb.create_sheet("Overview", 0)
    ws.column_dimensions["A"].width = 34

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    title_cell = ws.cell(row=1, column=1, value=f"Rental Property Deal Comparison — {now_str}")
    title_cell.font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(results) + 1)

    ws.cell(row=3, column=1, value="Metric").font = Font(bold=True)
    for i, r in enumerate(results, start=2):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = 20
        hdr = ws.cell(row=3, column=i, value=r.listing.address)
        hdr.font = Font(bold=True)
        hdr.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A4"

    def ins_label(r): return f"Insurance ({'OneHome' if r.listing.insurance_annual_listed else 'Assumed'})"
    def mnt_label(r): return f"Maintenance ({'OneHome' if r.listing.maintenance_annual_listed else 'Assumed'})"

    METRICS = [
        ("--- PROPERTY ---", None, None),
        ("List Price", "listing.list_price", CURRENCY_FMT),
        ("Property Type", "listing.property_type", None),
        ("Year Built", "listing.year_built", None),
        ("Beds / Baths", "_beds_baths", None),
        ("Sqft", "listing.sqft", NUMBER_FMT),
        ("Total Units", "listing.total_units", None),
        ("Days on Market", "listing.days_on_market", None),
        ("--- LOAN ---", None, None),
        ("Loan Amount", "loan_amount", CURRENCY_FMT),
        ("LTV", "ltv", PERCENT_FMT),
        ("Monthly P&I", "monthly_payment", CURRENCY_FMT),
        ("Total Cash Invested", "total_cash_invested", CURRENCY_FMT),
        ("--- INCOME ---", None, None),
        ("Rent Source", "rent_source", None),
        ("Rent Roll (annual)", "rent_roll_annual", CURRENCY_FMT),
        ("Listed Gross Income (OneHome)", "listed_gross_income", CURRENCY_FMT),
        ("Gross Income Used", "gross_rental_income", CURRENCY_FMT),
        ("Income Basis", "income_basis", None),
        ("Vacancy Loss", "vacancy_loss", CURRENCY_FMT),
        ("Effective Gross Income", "effective_gross_income", CURRENCY_FMT),
        ("--- EXPENSES ---", None, None),
        ("Property Taxes (OneHome)", "taxes_annual", CURRENCY_FMT),
        ("_insurance_label", "_insurance_val", CURRENCY_FMT),
        ("Property Management", "mgmt_fee_annual", CURRENCY_FMT),
        ("_maintenance_label", "_maintenance_val", CURRENCY_FMT),
        ("HOA (Annual)", "hoa_annual", CURRENCY_FMT),
        ("Electric (OneHome)", "electric_annual", CURRENCY_FMT),
        ("Utilities / Trash / Water / Sewer", "_utilities_group", CURRENCY_FMT),
        ("Other Listed Expenses (OneHome)", "misc_listed_expenses", CURRENCY_FMT),
        ("Operating Expenses (itemized)", "operating_expenses", CURRENCY_FMT),
        ("Operating Expenses (used)", "operating_expenses_used", CURRENCY_FMT),
        ("OpEx Basis", "opex_basis", None),
        ("CapEx Reserve", "capex_reserve", CURRENCY_FMT),
        ("Total Net Operating Expenses", "total_net_operating_expenses", CURRENCY_FMT),
        ("--- OPEX CHECK (vs listing) ---", None, None),
        ("Listed Operating Expense", "listed_operating_expenses", CURRENCY_FMT),
        ("Listed OpEx Source", "listed_opex_source", None),
        ("Calculated OpEx", "operating_expenses", CURRENCY_FMT),
        ("OpEx Variance (Calc − Listed)", "opex_variance", CURRENCY_FMT),
        ("OpEx Assessment", "_opex_flag", None),
        ("--- NOI & CASH FLOW ---", None, None),
        ("NOI — Calculated", "noi", CURRENCY_FMT),
        ("NOI — Listed (OneHome)", "listed_noi", CURRENCY_FMT),
        ("Effective NOI (after CapEx)", "effective_noi", CURRENCY_FMT),
        ("Annual Debt Service", "annual_debt_service", CURRENCY_FMT),
        ("Leveraged Cash Flow (Annual)", "cash_flow_annual", CURRENCY_FMT),
        ("Cash Flow (Monthly)", "cash_flow_monthly", CURRENCY_FMT),
        ("--- RETURNS ---", None, None),
        ("Cap Rate", "cap_rate", PERCENT_FMT),
        ("Cash-on-Cash (Annual)", "cash_on_cash", PERCENT_FMT),
        ("GRM", "grm", NUMBER_FMT),
        ("DSCR", "dscr", NUMBER_FMT),
        ("--- DECISION ---", None, None),
        ("Monthly Cash Flow", "cash_flow_monthly", CURRENCY_FMT),
        ("Target", "_target_cf", CURRENCY_FMT),
        ("Annual CoC Return", "cash_on_cash", PERCENT_FMT),
        ("Target", "_target_coc", PERCENT_FMT),
        ("Cap Rate", "cap_rate", PERCENT_FMT),
        ("Target", "_target_cr", PERCENT_FMT),
        ("DSCR", "dscr", NUMBER_FMT),
        ("Target", "_target_dscr", NUMBER_FMT),
        ("DEAL VERDICT", "_verdict", None),
    ]

    row = 4
    for label, attr, fmt in METRICS:
        if attr is None:
            for col in range(1, len(results) + 2):
                ws.cell(row=row, column=col).fill = SECTION_FILL
            ws.cell(row=row, column=1, value=label).font = Font(bold=True, color="1F4E79")
            row += 1
            continue

        if label == "_insurance_label":
            ws.cell(row=row, column=1, value=ins_label(results[0]))
        elif label == "_maintenance_label":
            ws.cell(row=row, column=1, value=mnt_label(results[0]))
        else:
            ws.cell(row=row, column=1, value=label)

        for i, result in enumerate(results, start=2):
            cell_val = None
            if attr == "_beds_baths":
                cell_val = f"{result.listing.beds}bd / {result.listing.baths}ba"
            elif attr == "_insurance_val":
                cell_val = result.insurance_annual
            elif attr == "_maintenance_val":
                cell_val = result.maintenance_annual
            elif attr == "_utilities_group":
                cell_val = (result.utilities_annual + result.trash_annual
                            + result.water_annual + result.sewer_annual + result.recycle_annual)
            elif attr == "_target_cf":
                cell_val = targets.min_monthly_cash_flow
            elif attr == "_target_coc":
                cell_val = targets.min_cash_on_cash_pct
            elif attr == "_target_cr":
                cell_val = targets.min_cap_rate_pct
            elif attr == "_target_dscr":
                cell_val = targets.min_dscr
            elif attr == "_opex_flag":
                if result.listed_operating_expenses is None:
                    cell_val = "N/A"
                else:
                    cell_val = "⚠ OVER" if result.opex_exceeds_listed else "✓ OK"
            elif attr == "_verdict":
                deal = evaluate_deal(result, targets)
                icons = {"GO": "🟢 GO", "BORDERLINE": "🟡 BORDERLINE", "NO-GO": "🔴 NO-GO"}
                cell_val = icons[deal["verdict"]]
            else:
                obj = result
                for part in attr.split("."):
                    obj = getattr(obj, part, None)
                    if obj is None:
                        break
                cell_val = obj

            cell = ws.cell(row=row, column=i, value=cell_val)
            if fmt and isinstance(cell_val, (int, float)):
                cell.number_format = fmt

            if attr in ("cash_flow_annual",) and isinstance(cell_val, (int, float)):
                cell.fill = GREEN_FILL if cell_val > 0 else (YELLOW_FILL if cell_val > -200 else RED_FILL)
            if attr == "cash_flow_monthly" and isinstance(cell_val, (int, float)):
                cell.fill = GREEN_FILL if cell_val > 0 else (YELLOW_FILL if cell_val > -17 else RED_FILL)

            if attr == "_opex_flag" and isinstance(cell_val, str):
                if "OVER" in cell_val:
                    cell.fill = RED_FILL
                    cell.font = Font(bold=True)
                elif "OK" in cell_val:
                    cell.fill = GREEN_FILL
                    cell.font = Font(bold=True)

            if attr == "_verdict" and isinstance(cell_val, str):
                if "🟢" in cell_val:
                    cell.fill = GREEN_FILL; cell.font = Font(bold=True)
                elif "🟡" in cell_val:
                    cell.fill = YELLOW_FILL; cell.font = Font(bold=True)
                else:
                    cell.fill = RED_FILL; cell.font = Font(bold=True, color="FFFFFF")

            if attr in ("_target_cf", "_target_coc", "_target_cr", "_target_dscr"):
                cell.fill = YELLOW_FILL
                cell.font = Font(italic=True, size=9)

        row += 1


# ── Main entry point ─────────────────────────────────────────────────────────

def build_workbook(results: list[AnalysisResult], output_path: Path, config: AnalysisConfig | None = None) -> Path:
    targets = config.targets if config else TargetConfig()
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    write_comparison_sheet(wb, results, targets)
    for result in results:
        write_property_sheet(wb, result, targets)

    wb.save(output_path)
    return output_path
